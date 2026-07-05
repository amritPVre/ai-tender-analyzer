"""PDF and Excel export utilities."""

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from config import COLORS


def _safe_str(value: Any, default: str = "Not specified") -> str:
    if value is None:
        return default
    if isinstance(value, (dict, list)):
        return str(value)
    s = str(value).strip()
    return s if s else default


def generate_pdf_report(
    doc_name: str,
    analysis: dict[str, Any],
) -> bytes:
    """Compile full analysis into a formatted PDF report."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor(COLORS["navy"]),
        spaceAfter=12,
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=colors.HexColor(COLORS["teal"]),
        spaceBefore=16,
        spaceAfter=8,
    )
    body_style = ParagraphStyle(
        "CustomBody",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
    )

    summary = analysis.get("summary", {})
    project_title = _safe_str(summary.get("project_title"), doc_name)

    story = [
        Paragraph("Solar Tender Intelligence Report", title_style),
        Paragraph(f"<b>Project:</b> {project_title}", body_style),
        Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d %B %Y, %H:%M')}", body_style),
        Paragraph(f"<b>Document:</b> {doc_name}", body_style),
        Spacer(1, 0.5 * cm),
    ]

    # Section 1 — Summary
    story.append(Paragraph("1. Document Summary", heading_style))
    summary_fields = [
        ("Project Title", summary.get("project_title")),
        ("Tendering Authority", summary.get("tendering_authority")),
        ("Capacity", summary.get("project_capacity")),
        ("Location", summary.get("project_location")),
        ("Technology", summary.get("technology_type")),
        ("Tender Type", summary.get("tender_type")),
        ("Estimated Value", summary.get("estimated_project_value")),
        ("Contract Duration", summary.get("contract_duration")),
    ]
    for label, val in summary_fields:
        story.append(Paragraph(f"<b>{label}:</b> {_safe_str(val)}", body_style))
    story.append(Paragraph(f"<b>Scope:</b> {_safe_str(summary.get('scope_summary'))}", body_style))

    # Section 2 — Technical
    story.append(Paragraph("2. Technical Requirements", heading_style))
    tech = _safe_str(analysis.get("technical", ""))
    for line in tech.replace("\r", "").split("\n"):
        if line.strip():
            story.append(Paragraph(line.replace("&", "&amp;"), body_style))

    # Section 3 — Procedure
    story.append(Paragraph("3. Tender Documentation Procedure", heading_style))
    proc = _safe_str(analysis.get("procedure", ""))
    for line in proc.replace("\r", "").split("\n"):
        if line.strip():
            story.append(Paragraph(line.replace("&", "&amp;"), body_style))

    # Section 4 — Checklist summary
    story.append(Paragraph("4. Document Submission Checklist", heading_style))
    checklist = analysis.get("checklist", {})
    if isinstance(checklist, dict):
        for category, items in checklist.items():
            if category == "raw_text":
                continue
            story.append(Paragraph(f"<b>{category}</b>", body_style))
            if isinstance(items, list):
                for item in items:
                    story.append(Paragraph(f"• {_safe_str(item)}", body_style))

    # Section 5 — Dates
    story.append(Paragraph("5. Key Dates & Milestones", heading_style))
    dates_data = analysis.get("dates", {})
    dates_list = dates_data.get("dates", []) if isinstance(dates_data, dict) else []
    for entry in dates_list:
        if isinstance(entry, dict):
            label = entry.get("label", "")
            date_val = entry.get("date", "")
            time_val = entry.get("time", "")
            venue = entry.get("venue", "")
            line = f"<b>{label}:</b> {_safe_str(date_val)}"
            if time_val and str(time_val).lower() not in ("not specified", ""):
                line += f" at {_safe_str(time_val)}"
            if venue and str(venue).lower() not in ("not specified", ""):
                line += f" — {_safe_str(venue)}"
            story.append(Paragraph(line, body_style))

    # Section 6 — Commercial
    story.append(Paragraph("6. Commercial & Risk Summary", heading_style))
    commercial_data = analysis.get("commercial", {})
    if isinstance(commercial_data, dict):
        comm = commercial_data.get("commercial", commercial_data)
        if isinstance(comm, dict):
            for key, val in comm.items():
                label = key.replace("_", " ").title()
                story.append(Paragraph(f"<b>{label}:</b> {_safe_str(val)}", body_style))

        risks = commercial_data.get("risks", [])
        if risks:
            story.append(Paragraph("<b>Risk Flags:</b>", body_style))
            for risk in risks:
                if isinstance(risk, dict):
                    story.append(
                        Paragraph(
                            f"[{risk.get('level', 'info').upper()}] "
                            f"{_safe_str(risk.get('title'))}: {_safe_str(risk.get('description'))}",
                            body_style,
                        )
                    )

        scoring = commercial_data.get("bid_scoring", {})
        if isinstance(scoring, dict):
            story.append(
                Paragraph(
                    f"<b>Recommendation:</b> {_safe_str(scoring.get('recommendation'))} "
                    f"(Score: {scoring.get('total_score', 'N/A')}/{scoring.get('max_score', 'N/A')})",
                    body_style,
                )
            )
            story.append(Paragraph(_safe_str(scoring.get("rationale")), body_style))

    # Section 7 — Eligibility
    story.append(Paragraph("7. Eligibility & Deviation Check", heading_style))
    elig_data = analysis.get("eligibility", {})
    if isinstance(elig_data, dict):
        for crit in elig_data.get("eligibility", []):
            if isinstance(crit, dict):
                story.append(
                    Paragraph(
                        f"• {_safe_str(crit.get('criterion'))}: {_safe_str(crit.get('requirement'))}",
                        body_style,
                    )
                )
        devs = elig_data.get("deviations", [])
        if devs:
            story.append(Paragraph("<b>Technical Deviations:</b>", body_style))
            for dev in devs:
                if isinstance(dev, dict):
                    story.append(
                        Paragraph(
                            f"• Spec: {_safe_str(dev.get('tender_spec'))} | "
                            f"Standard: {_safe_str(dev.get('applicable_standard'))} | "
                            f"Deviation: {_safe_str(dev.get('deviation_nature'))}",
                            body_style,
                        )
                    )

    def add_footer(canvas, doc_template):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor(COLORS["slate"]))
        canvas.drawCentredString(A4[0] / 2, 1 * cm, "Solar Tender Intelligence — Powered by BAESS.APP")
        canvas.drawRightString(A4[0] - 2 * cm, 1 * cm, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    buffer.seek(0)
    return buffer.read()


def generate_excel_checklist(checklist: dict[str, Any]) -> bytes:
    """Export checklist to Excel with one sheet per category."""
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="0B1F3A", end_color="0B1F3A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    status_fill = PatternFill(start_color="FAF3DC", end_color="FAF3DC", fill_type="solid")

    if not isinstance(checklist, dict) or not checklist:
        ws = wb.create_sheet("Checklist")
        ws.append(["Document", "Status"])
        ws["A1"].fill = header_fill
        ws["A1"].font = header_font
        ws["B1"].fill = header_fill
        ws["B1"].font = header_font
    else:
        for category, items in checklist.items():
            if category == "raw_text":
                continue
            sheet_name = str(category)[:31]
            ws = wb.create_sheet(sheet_name)
            ws.append(["Document / Item", "Prepared (Y/N)", "Notes"])
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            if isinstance(items, list):
                for item in items:
                    ws.append([str(item), "", ""])
                    for cell in ws[ws.max_row]:
                        cell.fill = status_fill

            ws.column_dimensions["A"].width = 60
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
